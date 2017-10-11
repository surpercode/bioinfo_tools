"""Mark the rows by patient and its id for patient gene mutation localtion"""

import openpyxl as opx
from openpyxl.styles import Font, Color, PatternFill, colors
import os
import re
import sys

def get_info_of_varified_gene(summary_xls_file, start_row = 3):
    """Extract the patient id and gene from summary_xls_file.
    Argv:
        summary_xls_file -- the summary of patient infos (1 col: id; 3 col: gene; 4 col: name)
    Return:
        info -- a dict {'varified': [(id, name, gene), ...],
                        'uncertain': [(id, name, gene), ...]}
    """
    wb = opx.load_workbook(summary_xls_file)
    sheet_names = wb.get_sheet_names()
    table = wb.get_sheet_by_name(sheet_names[0]) 
    nrows, ncols = table.max_row, table.max_column
    info = {'varified':[], 'uncertain':[]}
    for row in range(start_row, nrows): # start from 4 row
        row = str(row)
        pid, gene, name = table['A' + row].value, table['C' + row].value, table['D' + row].value
        pid = re.sub(r'[-_.]', '_', str(pid).strip()).split('_')[0]
        gene, name = gene.strip() if gene else gene, name.strip() if name else name
        #print(pid, gene, name)
        if gene and gene.strip() != '无':
            if gene.strip()[-1] == '?':
                info['uncertain'].append((str(pid), gene, name))
            else:
                gene = gene[:-1]
                info['varified'].append((str(pid), gene, name))
    return info

def read_patient_gene_loc(gene_loc_xls_file, start_row = 4):
    wb = opx.load_workbook(gene_loc_xls_file)
    sheet_names = wb.get_sheet_names()
    table = wb.get_sheet_by_name(sheet_names[1]) # sheet2
    nrows, ncols = table.max_row, table.max_column
    patient_gene = {}
    for row in range(start_row, nrows):
        row = str(row)
        pid, gene, geno, ref, alt = table['B' + row].value, table['E' + row].value, table['J' + row].value, table['K' + row].value, table['L' + row].value
        pid, gene, geno, ref, alt = [str(x).strip() for x in [pid, gene, geno, ref, alt]]
        try:
            patient_gene[(str(pid).strip(), gene)].append((row, str(geno).strip(), alt, ref))
        except KeyError as e:
            patient_gene[(str(pid).strip(), gene)] = []
            patient_gene[(str(pid), gene)].append((row, str(geno).strip(), alt, ref))
    return table, patient_gene
  
def mark_patient(patient_gene, info):
    varified_pids = set([x[0] for x in info['varified']])
    uncertain_pids = set([x[0] for x in info['uncertain']])
    patient_mark = {}
    for k, v in patient_gene.items():
        pid, gene = k
        mark = None
        try:
            pid_elements = re.sub(r'[-_.]', '-', str(pid)).split('-')
            if len(pid_elements) == 3:
                pid_elements.pop(1)
            #print(pid_elements)
            if set(pid_elements) & varified_pids:
                mark = 'varified'
            elif set(pid_elements) & uncertain_pids:
                mark = 'uncertain'
        except TypeError as e:
            pass
        patient_mark[pid] = mark
    return patient_mark

def patient_gene_select(patient_gene):
    patient_gene_match = set()
    for k, v in patient_gene.items():
        geno_match = {'0/1':[], '1/1':[], 'sp':[]}
        for row, geno, alt, ref in v:
            #print([type(x) for x in (row, geno, alt, ref)])
            try:
                int(alt)
            except ValueError as e:
                geno_match['sp'].append(row)
                continue
            if int(alt) >=5:
                if geno == '1/1':
                    geno_match[geno].append(row)
                elif geno == '0/1':
                    if int(ref) / int(alt) <=5:
                        geno_match[geno].append(row)
        if geno_match['1/1'] or len(geno_match['0/1']) >= 2 or geno_match['sp']:
            patient_gene_match.add(k)
            #print((row, geno, alt, ref),geno_match)
    return patient_gene_match

def read_gene_check(gene_check_file):
    pattern_gene = {0:set(), 1:set(), 2:set()}
    with open(gene_check_file, 'r') as f:
        for line in f:
            try:
                gene, pattern, name = line.strip().split('\t')
            except ValueError as e:
                continue
            pattern_gene[int(pattern)].add(gene.upper())
    return pattern_gene

def filter_gene(patient_gene, pattern_gene):
    patient_gene_list_left = set()
    pattern_to_filter = [0, 1]
    pattern_genes = set()
    for p in pattern_to_filter:
        pattern_genes |= pattern_gene[p]
    for k in patient_gene:
        pid, gene = k
        if not set(gene.upper().split(',')) & pattern_genes:
            patient_gene_list_left.add(k)
    return patient_gene_list_left

def copy_excel_cell(old_cell, new_cell):
    new_cell.value = old_cell.value
    if old_cell.has_style:
        try:
            new_cell.font = old_cell.font
            new_cell.border = old_cell.border
            new_cell.fill = old_cell.fill
            new_cell.number_format = old_cell.number_format
            new_cell.protection = old_cell.protection
            new_cell.alignment = old_cell.alignment
        except TypeError as e:
            pass

def write_excel_cell(old_cell, new_cell, font_style = None, fill_style = None):
    copy_excel_cell(old_cell, new_cell)
    if font_style:
        new_cell.font = font_style
    if fill_style:
        new_cell.fill = fill_style

def write_excel(file_name, table, patient_gene, patient_gene_list_left, patient_gene_match, patient_mark, gene_loc_start_row):
    print('first 10 patient gene match list is:\n', list(patient_gene_match)[:10])
    print('first 10 patient gene left list is:\n', list(patient_gene_list_left)[:10])
    patient_gene_good = patient_gene_list_left & patient_gene_match
    print('patient_gene_good size is:', len(patient_gene_good))
    max_row, max_col, min_row, min_col = table.max_row, table.max_column, table.min_row, table.min_column
    red_font = Font(bold = True, color = colors.RED)
    blue_font = Font(bold = True, color = colors.BLUE)
    yellow_fill = PatternFill(fill_type = 'solid', start_color = colors.YELLOW, end_color = colors.YELLOW)
    good_row_styles = []
    for x in range(1, gene_loc_start_row):
        good_row_styles.append((x, None, None))
    gene_count = {}
    gene_patient_count = {}
    row_gene = {}
    for k in patient_gene_good:
        pid, gene = k
        try:
            gene_count[gene.upper()] += len(patient_gene[k])
            gene_patient_count[gene.upper()] += 1
        except KeyError as e:
            gene_count[gene.upper()] = len(patient_gene[k])
            gene_patient_count[gene.upper()] = 1
        font_style, fill_style = None, None
        if patient_mark[pid] == 'varified':
            font_style = red_font
            fill_style = yellow_fill
        elif patient_mark[pid] == 'uncertain':
            font_style = blue_font
            fill_style = yellow_fill
        for t in patient_gene[k]:
            row = int(t[0])
            row_gene[row] = gene.upper()
            good_row_styles.append((row, font_style, fill_style))
    good_row_styles = sorted(good_row_styles, key = lambda x:x[0])
    wb = opx.Workbook()
    new_table = wb.active
    for i, term in enumerate(good_row_styles):
        row, font_style, fill_style = term
        for col in range(min_col, max_col + 1):
            if col == 6 and row >= gene_loc_start_row :
                new_table.cell(row = i + 1, column = col).value = str(gene_count[row_gene[row]]) + '/' + str(gene_patient_count[row_gene[row]])
                continue
            new_col, old_col = col, col
            if col > 6:
                new_col += 1
            old_cell, new_cell = table.cell(row = row, column = old_col), new_table.cell(row = i + 1, column = new_col)
            write_excel_cell(old_cell, new_cell, font_style, fill_style)
    wb.save(file_name)
    return patient_gene_good

def main(summary_xls_file, gene_loc_xls_file, gene_check_file, output_file, summary_start_row, gene_loc_start_row):
    summary_info = get_info_of_varified_gene(summary_xls_file, summary_start_row)
    table, patient_gene = read_patient_gene_loc(gene_loc_xls_file, gene_loc_start_row)
    print('patient_gene size is:', len(patient_gene))
    patient_gene_mark = mark_patient(patient_gene, summary_info)
    print('paitient_gene_mark size is:', len(patient_gene_mark))
    patient_gene_match = patient_gene_select(patient_gene)
    print('patient_gene_match size is:', len(patient_gene_match))
    gene_check_list = read_gene_check(gene_check_file)
    print('gene_check_list size is:', len(gene_check_list))
    patient_gene_list_left = filter_gene(patient_gene, gene_check_list)
    print('patient_gene_list_left size is:', len(patient_gene_list_left))
    patient_gene_good = write_excel(output_file, table, patient_gene, patient_gene_list_left, patient_gene_match, patient_gene_mark, gene_loc_start_row)
    return patient_gene_good

if __name__ == '__main__':
    summary_xls_file = '../summary.xlsx'
    gene_loc_xls_file = '../chun_fu_to_be_varified_2.xlsx'
    #gene_loc_xls_file = '../chun_fu_to_be_varified.xlsx'
    gene_check_file = '../../gene_check_wenjing/gene_check_table_modified.txt'
    output_file = '.'.join(gene_loc_xls_file.split('.')[:-1]) + '_filter_and_mark.xlsx'
    print(output_file)
    patient_gene_good = main(summary_xls_file, gene_loc_xls_file, gene_check_file, output_file, 3, 1)
    #patient_gene_good = main(summary_xls_file, gene_loc_xls_file, gene_check_file, output_file, 3, 4)
    patient_gene_good_out = '.'.join(gene_loc_xls_file.split('.')[:-1]) + '_good.txt'
    with open(patient_gene_good_out, 'w') as fo:
        for t in patient_gene_good:
            fo.write('\t'.join(t) + '\n')
